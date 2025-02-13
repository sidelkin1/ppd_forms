from datetime import date
from pathlib import Path
from shutil import make_archive
from typing import NamedTuple, cast

import openpyxl
import pandas as pd
from openpyxl.cell import Cell
from openpyxl.chart import LineChart, Reference
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import save_workbook

from app.core.models.dto import UneftFieldDB, UneftReservoirDB
from app.core.utils.process_pool import ProcessPoolManager
from app.infrastructure.db.dao.complex.reporters import MatbalReporter

_CELL_ROW_NUM = "A15"
_CELL_DATE = "B15"
_CELL_OIL_RATE = "C15"
_CELL_WAT_RATE = "D15"
_CELL_INJ_RATE = "E15"
_CELL_ACTUAL_RESP = "F15"
_CELL_RESP_WEIGHT = "G15"
_CELL_CALC_RESP = "H15"
_CELL_AQ_RATE = "J15"
_CELL_BOUND_RATE = "K15"
_CELL_DIFF_RESP = "L15"
_CELL_RHS_TERM_1 = "M15"
_CELL_RHS_TERM_2 = "N15"
_CELL_CALC_LIQ_RATE = "O15"
_CELL_CALC_INJ_RATE = "P15"

_RANGE_CALC_RESP = "H15:I15"

_FORMULA_CALC_RESP = Translator(
    "=TRANSPOSE(MMULT($U$11:$V$12,TRANSPOSE(M15:N15)))", _CELL_CALC_RESP
)
_FORMULA_AQ_RATE = Translator("=alpha*(I15-H15)", _CELL_AQ_RATE)
_FORMULA_BOUND_RATE = Translator("=beta*(Pi-I15)", _CELL_BOUND_RATE)
_FORMULA_DIFF_RESP = Translator(
    '=IF(F15>0,G15*(H15-F15)^2," ")', _CELL_DIFF_RESP
)
_FORMULA_RHS_TERM_1 = Translator(
    "=N*1000*Boi*Cfactor*H14"
    "+(lambda*E15-D15)*B(Bwi,Pi,Cw,H14)"
    "-C15*B(Boi,Pi,Co,H14)",
    _CELL_RHS_TERM_1,
)
_FORMULA_RHS_TERM_2 = Translator(
    "=Waq*1000*(Cw+Cf)*I14+beta*Pi", _CELL_RHS_TERM_2
)
_FORMULA_CALC_LIQ_RATE = Translator(
    "=(C15+D15)/(EDATE(B15,1)-B15)", _CELL_CALC_LIQ_RATE
)
_FORMULA_CALC_INJ_RATE = Translator(
    "=E15/(EDATE(B15,1)-B15)*lambda", _CELL_CALC_INJ_RATE
)


class _DataFrameRow(NamedTuple):
    date: date
    Qoil: float
    Qwat: float
    Qinj: float
    Pres: float


def _expand_date_range(
    df: pd.DataFrame, left: date, right: date, result: str
) -> pd.DataFrame:
    df[result] = pd.date_range(start=left, end=right, freq="MS").date
    df = df.explode(result)
    return df


def _shift_to_month_start(s: pd.Series) -> pd.Series:
    s = pd.to_datetime(s, errors="coerce")
    crit = s.dt.is_month_start
    s[~crit] -= pd.offsets.MonthBegin(n=1)  # type: ignore[operator]
    return s.dt.date


def _prepare_measurements(df: pd.DataFrame) -> pd.DataFrame:
    df["date"] = _shift_to_month_start(df["date"])
    df = df.groupby("date", as_index=False).mean()
    return df


def _join_rates_and_measurements(
    df: pd.DataFrame, rates: pd.DataFrame, measurements: pd.DataFrame | None
) -> pd.DataFrame:
    df = pd.merge(df, rates, how="left", on="date").fillna(0)
    if measurements is not None:
        measurements = _prepare_measurements(measurements)
        df = pd.merge(df, measurements, how="left", on="date")
    else:
        df = df.assign(Pres=None)
    return df


def _fill_rates_and_measurements(
    ws: Worksheet, row_num: int, df_row: _DataFrameRow
) -> None:
    ws[_CELL_ROW_NUM].offset(row=row_num).value = row_num + 1
    ws[_CELL_DATE].offset(row=row_num).value = df_row.date
    ws[_CELL_OIL_RATE].offset(row=row_num).value = df_row.Qoil
    ws[_CELL_WAT_RATE].offset(row=row_num).value = df_row.Qwat
    ws[_CELL_INJ_RATE].offset(row=row_num).value = df_row.Qinj
    ws[_CELL_ACTUAL_RESP].offset(row=row_num).value = df_row.Pres
    ws[_CELL_RESP_WEIGHT].offset(row=row_num).value = 1 if df_row.Pres else ""


def _fill_formulas(ws: Worksheet, row_num: int, calc_resp_range: str) -> None:
    ws[_CELL_CALC_RESP].offset(row=row_num).value = ArrayFormula(
        calc_resp_range,
        _FORMULA_CALC_RESP.translate_formula(row_delta=row_num),
    )
    ws[_CELL_AQ_RATE].offset(
        row=row_num
    ).value = _FORMULA_AQ_RATE.translate_formula(row_delta=row_num)
    ws[_CELL_BOUND_RATE].offset(
        row=row_num
    ).value = _FORMULA_BOUND_RATE.translate_formula(row_delta=row_num)
    ws[_CELL_DIFF_RESP].offset(
        row=row_num
    ).value = _FORMULA_DIFF_RESP.translate_formula(row_delta=row_num)
    ws[_CELL_RHS_TERM_1].offset(
        row=row_num
    ).value = _FORMULA_RHS_TERM_1.translate_formula(row_delta=row_num)
    ws[_CELL_RHS_TERM_2].offset(
        row=row_num
    ).value = _FORMULA_RHS_TERM_2.translate_formula(row_delta=row_num)
    ws[_CELL_CALC_LIQ_RATE].offset(
        row=row_num
    ).value = _FORMULA_CALC_LIQ_RATE.translate_formula(row_delta=row_num)
    ws[_CELL_CALC_INJ_RATE].offset(
        row=row_num
    ).value = _FORMULA_CALC_INJ_RATE.translate_formula(row_delta=row_num)


def _edit_charts(ws: Worksheet, nrow: int) -> None:
    cell = cast(Cell, ws[_CELL_DATE])
    dates = Reference(
        ws, min_col=cell.column, min_row=cell.row, max_row=cell.row + nrow - 1
    )
    for chart in ws._charts:  # type: ignore[attr-defined]
        chart = cast(LineChart, chart)
        chart.set_categories(dates)


def _fill_template(df: pd.DataFrame, path: Path, template: Path) -> None:
    result = path / template.name
    try:
        wb = openpyxl.load_workbook(template, keep_vba=True)
        ws = wb["MB_simple"]
        calc_range = CellRange(_RANGE_CALC_RESP)
        for row_num, df_row in enumerate(df.itertuples(index=False)):
            _fill_rates_and_measurements(
                ws, row_num, cast(_DataFrameRow, df_row)
            )
            _fill_formulas(ws, row_num, calc_range.coord)
            calc_range.shift(row_shift=1)
        _edit_charts(ws, df.shape[0])
        save_workbook(wb, result)
    finally:
        wb.close()


def _process_data(
    rates: pd.DataFrame,
    measurements: pd.DataFrame | None,
    path: Path,
    template: Path,
) -> pd.DataFrame:
    df = pd.DataFrame(columns=["date"])
    df = _expand_date_range(
        df, rates["date"].min(), rates["date"].max(), "date"
    )
    df = _join_rates_and_measurements(df, rates, measurements)
    _fill_template(df, path, template)
    df.to_csv(path / "matbal.csv", sep=";", date_format="%d.%m.%Y")
    return df


async def matbal_report(
    path: Path,
    template: Path,
    field: UneftFieldDB,
    reservoirs: list[UneftReservoirDB],
    alternative: bool,
    dao: MatbalReporter,
    pool: ProcessPoolManager,
) -> None:
    reservoir_names = [reservoir.name for reservoir in reservoirs]
    rates = await dao.get_rates(field.id, reservoir_names, alternative)
    measurements = await dao.get_measurements()
    await pool.run(_process_data, rates, measurements, path, template)
    make_archive(str(path), "zip", root_dir=path)
