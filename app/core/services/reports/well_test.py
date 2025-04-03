from copy import copy, deepcopy
from datetime import date
from io import BytesIO
from pathlib import Path
from shutil import make_archive
from typing import Any, NamedTuple, cast

import openpyxl
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
from openpyxl.styles import Border, Side
from openpyxl.styles.fills import FILL_SOLID, PatternFill
from openpyxl.utils import quote_sheetname
from openpyxl.utils.datetime import to_excel
from openpyxl.utils.inference import cast_numeric
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import save_workbook
from PIL import Image as PILImage
from PIL import ImageOps

from app.core.models.dto import WellTestResult
from app.core.utils.process_pool import ProcessPoolManager
from app.infrastructure.db.dao.complex.reporters import WellTestReporter

_EXCEL_GDIS_START_ROW = 2
_EXCEL_GDIS_START_COLUMN = 1
_EXCEL_GDIS_LAST_COLUMN = 12

_EXCEL_GTM_START_ROW = 2
_EXCEL_GTM_START_COLUMN = 14
_EXCEL_GTM_LAST_COLUMN = 20

_EXCEL_NEIGHBS_START_ROW = 2
_EXCEL_NEIGHBS_START_COLUMN = 22
_EXCEL_NEIGHBS_LAST_COLUMN = 29

_REPORT_SHEET_NAME = "отчет"

_REPORT_ISOBARS_SHIFT = 3
_REPORT_ISOBARS_WIDTH_MM = 91.0
_REPORT_ISOBARS_HEIGHT_MM = 75.7
_REPORT_ISOBARS_DPI = 96
_REPORT_ISOBARS_WIDTH_PX = int(
    round(_REPORT_ISOBARS_WIDTH_MM * _REPORT_ISOBARS_DPI / 25.4)
)
_REPORT_ISOBARS_HEIGHT_PX = int(
    round(_REPORT_ISOBARS_HEIGHT_MM * _REPORT_ISOBARS_DPI / 25.4)
)

_REPORT_CELL_TITLE = "B3"
_REPORT_CELL_DATE = "E4"
_REPORT_CELL_RESERVOIR = "E5"
_REPORT_CELL_PURPOSE = "E6"
_REPORT_CELL_TEXT = "B7"
_REPORT_CELL_CONCLUSION = "B8"
_REPORT_CELL_TEST_WELL = "B9"
_REPORT_CELL_TEST_NAME = "C9"
_REPORT_CELL_TEST_DATE = "D9"
_REPORT_CELL_TEST_PRESSURE = "E9"
_REPORT_CELL_TEST_PRESSURE_PLOT = "M9"
_REPORT_CELL_TEST_PERMEABILITY = "F9"
_REPORT_CELL_TEST_SKIN_FACTOR = "G9"
_REPORT_CELL_TEST_PROD_INDEX = "H9"
_REPORT_CELL_TEST_FRAC_LENGTH = "I9"
_REPORT_CELL_TEST_RELIABILITY = "J9"
_REPORT_RANGE_GTM_NAME = (
    f"{_REPORT_CELL_TEST_PRESSURE}:{_REPORT_CELL_TEST_RELIABILITY}"
)

_REPORT_CHART_SHIFT = 2
_REPORT_CHART_HEIGHT = 15
_REPORT_CHART_WIDTH = 4
_REPORT_CHART_INDEX = 0

_REPORT_FOOTER_SHIFT = 18
_REPORT_FOOTER_INDEX = 3

_REPORT_ARROW_SHIFT = 7

_REPORT_ROW_CELL_BORDER = Border(
    left=Side(border_style="thin", color="000066CC"),
    right=Side(border_style="thin", color="000066CC"),
    top=Side(border_style="thin", color="000066CC"),
    bottom=Side(border_style="thin", color="000066CC"),
)
_REPORT_ROW_CELL_FILL = PatternFill(FILL_SOLID, fgColor="F2F2F2")

_REPORT_TABLE_ROW_CELLS = (
    {"cell": _REPORT_CELL_TEST_WELL, "value": "well"},
    {"cell": _REPORT_CELL_TEST_NAME, "value": "well_test"},
    {"cell": _REPORT_CELL_TEST_DATE, "value": "end_date"},
    {"cell": _REPORT_CELL_TEST_PRESSURE, "value": "resp_owc"},
    {"cell": _REPORT_CELL_TEST_PRESSURE_PLOT, "value": "resp_owc"},
    {"cell": _REPORT_CELL_TEST_PERMEABILITY, "value": "permeability"},
    {"cell": _REPORT_CELL_TEST_SKIN_FACTOR, "value": "skin_factor"},
    {"cell": _REPORT_CELL_TEST_PROD_INDEX, "value": "prod_index"},
    {"cell": _REPORT_CELL_TEST_FRAC_LENGTH, "value": "frac_length"},
    {"cell": _REPORT_CELL_TEST_RELIABILITY, "value": "reliability"},
)
_REPORT_TABLE_ROW_CELLS_GTM = (
    {"cell": _REPORT_CELL_TEST_WELL, "value": "well"},
    {"cell": _REPORT_CELL_TEST_NAME, "value": "well_test"},
    {"cell": _REPORT_CELL_TEST_DATE, "value": "end_date"},
    {"cell": _REPORT_CELL_TEST_PRESSURE, "value": "reliability"},
    {"cell": _REPORT_CELL_TEST_PRESSURE_PLOT, "value": None},
    {"cell": _REPORT_CELL_TEST_PERMEABILITY, "value": None},
    {"cell": _REPORT_CELL_TEST_SKIN_FACTOR, "value": None},
    {"cell": _REPORT_CELL_TEST_PROD_INDEX, "value": None},
    {"cell": _REPORT_CELL_TEST_FRAC_LENGTH, "value": None},
    {"cell": _REPORT_CELL_TEST_RELIABILITY, "value": None},
)


class _ReportTableRow(NamedTuple):
    well: str
    well_test: str
    end_date: date
    resp_owc: float
    permeability: float
    skin_factor: float
    prod_index: float
    frac_length: float
    reliability: str
    is_gtm: bool


def _get_uids(neighbs: pd.DataFrame) -> list[str]:
    if neighbs.empty:
        return []
    cols = ["field", "well", "reservoir"]
    return neighbs[cols].agg("".join, axis=1).to_list()


def _fill_template(
    ws: Worksheet, df: pd.DataFrame, min_row: int, min_col: int, max_col: int
) -> None:
    for row, df_row in zip(
        ws.iter_rows(
            min_row=min_row,
            min_col=min_col,
            max_col=max_col,
            max_row=df.shape[0] + min_row - 1,
        ),
        df.itertuples(index=False),
    ):
        for cell, value in zip(row, df_row):
            cell.value = value


def _concat_tests(
    tests: pd.DataFrame, results: list[WellTestResult]
) -> pd.DataFrame:
    df = pd.DataFrame(results, columns=tests.columns)
    return df if tests.empty else pd.concat([tests, df], ignore_index=True)


def _add_distance(tests: pd.DataFrame, neighbs: pd.DataFrame) -> pd.DataFrame:
    cols = ["field", "well", "reservoir"]
    df = pd.merge(tests, neighbs, on=cols, how="left").sort_values(
        ["reservoir", "distance"]
    )
    df["well"] = df["well_layer"]
    return df.drop(columns="well_layer")


def _resize_isobars(isobars: Image) -> Image:
    image_stream = cast(BytesIO, isobars.ref)
    image_stream.seek(0)
    with PILImage.open(image_stream) as pil_image:
        pil_image.load()
        resized_image = ImageOps.fit(
            pil_image,
            (_REPORT_ISOBARS_WIDTH_PX, _REPORT_ISOBARS_HEIGHT_PX),
        )
    output_stream = BytesIO()
    resized_image.save(output_stream, format=isobars.format)
    output_stream.seek(0)
    return Image(output_stream)


def _fill_data_sheet(
    wb: openpyxl.Workbook,
    gtms: pd.DataFrame,
    tests: pd.DataFrame,
    neighb_tests: pd.DataFrame,
) -> None:
    ws = wb["Лист1"]
    start_row = _EXCEL_GDIS_START_ROW
    for _, df in tests.groupby("reservoir"):
        _fill_template(
            ws,
            df,
            start_row,
            _EXCEL_GDIS_START_COLUMN,
            _EXCEL_GDIS_LAST_COLUMN,
        )
        start_row += df.shape[0]
    _fill_template(
        ws,
        gtms,
        _EXCEL_GTM_START_ROW,
        _EXCEL_GTM_START_COLUMN,
        _EXCEL_GTM_LAST_COLUMN,
    )
    _fill_template(
        ws,
        neighb_tests,
        _EXCEL_NEIGHBS_START_ROW,
        _EXCEL_NEIGHBS_START_COLUMN,
        _EXCEL_NEIGHBS_LAST_COLUMN,
    )


def _copy_sheets(
    wb: openpyxl.Workbook, reservoirs: list[str]
) -> dict[str, Worksheet]:
    ws = wb[_REPORT_SHEET_NAME]
    if len(reservoirs) == 1:
        return {reservoirs[0]: ws}
    sheets = {}
    for reservoir in reservoirs:
        copy_ws = wb.copy_worksheet(ws)
        copy_ws.title = f"{_REPORT_SHEET_NAME} ({reservoir})"
        copy_ws.views = copy(ws.views)
        copy_ws._images = deepcopy(ws._images)  # type: ignore[attr-defined]
        copy_ws._charts = deepcopy(ws._charts)  # type: ignore[attr-defined]
        copy_ws.sheet_view.tabSelected = False
        sheets[reservoir] = copy_ws
    wb.remove(ws)
    return sheets


def _concat_gtms(tests: pd.DataFrame, gtms: pd.DataFrame) -> pd.DataFrame:
    tests = tests.assign(is_gtm=False)
    if not gtms.empty:
        df = pd.DataFrame(
            {
                "field": gtms["field"],
                "well": gtms["well"],
                "reservoir": gtms["reservoir"],
                "well_type": gtms["well_type"],
                "well_test": "ГТМ",
                "end_date": gtms["gtm_date"],
                "reliability": gtms["gtm_description"].fillna(
                    gtms["gtm_name"]
                ),
                "is_gtm": True,
            }
        )
        tests = pd.concat([tests, df]).sort_values("end_date")
    # Для ячеек с номером скважины, состоящим из одних цифр,
    # Excel вставляет сообщение, что "Число сохранено как текст",
    # поэтому специально преобразуем номер в `int`
    tests.loc[:, "well"] = tests["well"].map(lambda x: cast_numeric(x) or x)
    return tests


def _fill_table_row(
    ws: Worksheet,
    df_row: _ReportTableRow,
    row_num: int,
    row_cells: tuple[dict[str, Any], ...],
) -> None:
    for row_cell in row_cells:
        cell = ws[row_cell["cell"]].offset(row=row_num)
        if row_cell["value"] is not None:
            cell.value = getattr(df_row, row_cell["value"])
        cell.border = _REPORT_ROW_CELL_BORDER
        if df_row.is_gtm:
            cell.fill = _REPORT_ROW_CELL_FILL


def _fill_test_history(ws: Worksheet, tests: pd.DataFrame) -> None:
    cell_range = CellRange(_REPORT_RANGE_GTM_NAME)
    for row_num, df_row in enumerate(tests.itertuples(index=False), start=1):
        cell_range.shift(row_shift=1)
        if df_row.is_gtm:
            _fill_table_row(
                ws,
                cast(_ReportTableRow, df_row),
                row_num,
                _REPORT_TABLE_ROW_CELLS_GTM,
            )
            ws.merge_cells(range_string=cell_range.coord)
        else:
            _fill_table_row(
                ws,
                cast(_ReportTableRow, df_row),
                row_num,
                _REPORT_TABLE_ROW_CELLS,
            )


def _shift_drawings(ws: Worksheet, nrow: int) -> None:
    chart = ws._charts[_REPORT_CHART_INDEX]  # type: ignore[attr-defined]
    cell = ws[_REPORT_CELL_TEST_PERMEABILITY].offset(
        _REPORT_CHART_SHIFT + nrow - 1
    )
    anchor = TwoCellAnchor()
    anchor._from.col = cell.column  # type: ignore[union-attr]
    anchor._from.row = cell.row  # type: ignore[union-attr]
    anchor.to.col = (  # type: ignore[union-attr]
        _REPORT_CHART_WIDTH + cell.column
    )
    anchor.to.row = (  # type: ignore[union-attr]
        _REPORT_CHART_HEIGHT + cell.row - 1
    )
    chart.anchor = anchor  # type: ignore[assignment]
    image = ws._images[_REPORT_FOOTER_INDEX]  # type: ignore[attr-defined]
    cell = ws[_REPORT_CELL_TEST_SKIN_FACTOR].offset(
        _REPORT_FOOTER_SHIFT + nrow - 1
    )
    image.anchor = cell.coordinate


def _edit_chart(
    ws: Worksheet,
    min_date: date,
    max_date: date,
    p_init: float,
    p_bubble: float,
) -> None:
    chart = ws._charts[_REPORT_CHART_INDEX]  # type: ignore[attr-defined]
    chart.x_axis.scaling.min = to_excel(min_date.replace(day=1))
    chart.x_axis.scaling.max = to_excel(
        max_date.replace(day=1) + relativedelta(months=1)
    )
    for series in chart.series:
        if series.title.value == "Pнач":
            series.yVal.numLit.pt[0].v = p_init
            series.yVal.numLit.pt[1].v = p_init
        elif series.title.value == "Pнас":
            series.yVal.numLit.pt[0].v = p_bubble
            series.yVal.numLit.pt[1].v = p_bubble
        else:
            for ref in (series.xVal.numRef, series.yVal.numRef):
                if ref is not None:
                    ref.f = ref.f.replace(
                        _REPORT_SHEET_NAME, quote_sheetname(ws.title)
                    )


def _process_drawings(
    ws: Worksheet,
    tests: pd.DataFrame,
    pvt: pd.DataFrame,
    isobars: Image | None,
    arrow: Path,
) -> None:
    _edit_chart(
        ws,
        tests.loc[~tests["is_gtm"], "end_date"].min(),
        tests.loc[~tests["is_gtm"], "end_date"].max(),
        pvt["p_init"].iat[0],
        pvt["p_bubble"].iat[0],
    )
    _shift_drawings(ws, tests.shape[0])
    if isobars is not None:
        resized_isobars = _resize_isobars(isobars)
        cell = ws[_REPORT_CELL_TEST_WELL].offset(
            _REPORT_ISOBARS_SHIFT + tests.shape[0] - 1
        )
        ws.add_image(resized_isobars, cell.coordinate)
        image = Image(arrow)
        cell = ws[_REPORT_CELL_TEST_NAME].offset(
            _REPORT_ARROW_SHIFT + tests.shape[0] - 1
        )
        ws.add_image(image, cell.coordinate)


def _fill_report_sheet(
    wb: openpyxl.Workbook,
    results: list[WellTestResult],
    tests: pd.DataFrame,
    pvt: pd.DataFrame,
    gtms: pd.DataFrame,
    arrow: Path,
) -> None:
    reservoirs = [result["reservoir"] for result in results]
    sheets = _copy_sheets(wb, reservoirs)
    for result in results:
        reservoir = result["reservoir"]
        ws = sheets[reservoir]
        test_group = tests[tests["reservoir"] == reservoir]
        pvt_group = pvt[pvt["reservoir"] == reservoir]
        gtm_group = gtms[
            gtms["reservoir"].str.contains(reservoir, regex=False)
            & (gtms["gtm_date"] > test_group["end_date"].min())
        ]
        test_date = result["end_date"].strftime("%d.%m.%Y")
        ws[_REPORT_CELL_TITLE].value = (
            f"Результаты"
            f" {result['well_test']}"
            f" {result['well']}"
            f" {result['field']}"
            f"\n{test_date}"
        )
        ws[_REPORT_CELL_DATE].value = test_date
        ws[_REPORT_CELL_RESERVOIR].value = result["report_reservoir"]
        ws[_REPORT_CELL_PURPOSE].value = result["purpose"]
        test_group = _concat_gtms(test_group, gtm_group)
        _fill_test_history(ws, test_group)
        _process_drawings(ws, test_group, pvt_group, result["isobars"], arrow)


def _process_data(
    results: list[WellTestResult],
    dfs: dict[str, pd.DataFrame],
    path: Path,
    template: Path,
    arrow: Path,
) -> None:
    result = path / template.name
    tests = _concat_tests(dfs["tests"], results)
    neighb_tests = _add_distance(dfs["neighb_tests"], dfs["neighbs"])
    try:
        wb = openpyxl.load_workbook(template)
        _fill_data_sheet(wb, dfs["gtms"], tests, neighb_tests)
        _fill_report_sheet(wb, results, tests, dfs["pvt"], dfs["gtms"], arrow)
        save_workbook(wb, result)
    finally:
        wb.close()


async def well_test_report(
    path: Path,
    template: Path,
    arrow: Path,
    gtm_period: int,
    gdis_period: int,
    radius: float,
    dao: WellTestReporter,
    pool: ProcessPoolManager,
) -> None:
    results = await dao.get_results()
    dfs = await dao.read_all(
        results=results,
        gtm_period=gtm_period,
        gdis_period=gdis_period,
        radius=radius,
    )
    await pool.run(_process_data, results, dfs, path, template, arrow)
    make_archive(str(path), "zip", root_dir=path)
