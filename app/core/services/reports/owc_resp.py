import asyncio
from datetime import date
from pathlib import Path
from shutil import make_archive

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import save_workbook

from app.core.models.dto.db.field_list import UneftFieldDB
from app.core.models.dto.db.reservoir_list import UneftReservoirDB
from app.core.models.enums import WellTest
from app.core.utils.process_pool import ProcessPoolManager
from app.infrastructure.db.dao.sql.reporters import OwcRespReporter

_CALCULATOR_FIELD = "B1"
_CALCULATOR_RESERVOIR = "B2"
_CALCULATOR_WELL_NAME = "B3"
_CALCULATOR_WELL_MODE = "B4"
_CALCULATOR_ELEVATION = "B5"
_CALCULATOR_OWC_ABS_DEPTH = "B6"
_CALCULATOR_TOP_PERF_DEPTH = "B7"
_CALCULATOR_MEASURED_PRESSURE = "B8"
_CALCULATOR_MEASURED_DEPTH = "B9"
_CALCULATOR_OIL_DENSITY = "B12"
_CALCULATOR_WATER_DENSITY = "B13"
_CALCULATOR_WATERCUT = "B14"
_CALCULATOR_LIQUID_DENSITY = "B15"

_ANALYTICS_REGION = "B5"
_ANALYTICS_WORKSHOP = "C5"
_ANALYTICS_FIELD = "E5"
_ANALYTICS_WELL_NAME = "F5"
_ANALYTICS_WELL_PAD = "G5"
_ANALYTICS_RESERVOIR = "I5"
_ANALYTICS_WELL_MODE = "J5"
_ANALYTICS_WELLBORE = "K5"
_ANALYTICS_WELL_LIFT = "L5"
_ANALYTICS_TOP_PERF_DEPTH = "M5"
_ANALYTICS_TOP_PERF_OFFSET = "N5"
_ANALYTICS_WATERCUT = "U5"
_ANALYTICS_WELL_STATUS = "V5"
_ANALYTICS_OIL_DENSITY = "AL5"
_ANALYTICS_WATER_DENSITY = "AM5"
_ANALYTICS_WELL_TEST_PLAN = "AW5"
_ANALYTICS_WELL_TEST_ACTUAL = "BD5"
_ANALYTICS_WELL_TEST_START_DATE = "BE5"
_ANALYTICS_WELL_TEST_END_DATE = "BF5"
_ANALYTICS_STATIC_LEVEL_DEPTH = "BL5"
_ANALYTICS_ANNULAR_PRESSURE = "BM5"
_ANALYTICS_MEASURED_DEPTH = "BN5"
_ANALYTICS_MEASURED_PRESSURE = "BO5"
_ANALYTICS_TOP_PERF_PRESSURE = "BY5"
_ANALYTICS_OWC_PRESSURE = "BZ5"


def _fill_calculator(ws: Worksheet, props: pd.DataFrame) -> None:
    ws[_CALCULATOR_FIELD].value = props["field"].item()
    ws[_CALCULATOR_RESERVOIR].value = props["reservoir"].item()
    ws[_CALCULATOR_WELL_NAME].value = props["branch"].item()
    ws[_CALCULATOR_WELL_MODE].value = props["well_mode"].item()
    ws[_CALCULATOR_ELEVATION].value = props["elevation"].item()
    ws[_CALCULATOR_OWC_ABS_DEPTH].value = props["abs_depth_owc"].item()
    ws[_CALCULATOR_TOP_PERF_DEPTH].value = props["top_perf"].item()
    ws[_CALCULATOR_MEASURED_PRESSURE].value = props["pressure"].item()
    ws[_CALCULATOR_MEASURED_DEPTH].value = props["depth"].item()
    ws[_CALCULATOR_OIL_DENSITY].value = props["layer_oil_density"].item()
    ws[_CALCULATOR_WATER_DENSITY].value = props["water_density"].item()
    ws[_CALCULATOR_WATERCUT].value = props["watercut"].item()
    ws[_CALCULATOR_LIQUID_DENSITY].value = props["liquid_density"].item()


def _fill_depth(ws: Worksheet, depths: pd.DataFrame) -> None:
    for row in dataframe_to_rows(depths, index=False, header=False):
        ws.append(row)


def _fill_analytics(
    ws: Worksheet, props: pd.DataFrame, well_test: WellTest
) -> None:
    on_date = pd.to_datetime(props["on_date"]).dt.strftime("%d.%m.%Y").item()
    ws[_ANALYTICS_REGION].value = props["region"].item()
    ws[_ANALYTICS_WORKSHOP].value = props["workshop"].item()
    ws[_ANALYTICS_FIELD].value = props["field"].item()
    ws[_ANALYTICS_WELL_NAME].value = props["well"].item()
    ws[_ANALYTICS_WELL_PAD].value = props["pad"].item()
    ws[_ANALYTICS_RESERVOIR].value = props["reservoir"].item()
    ws[_ANALYTICS_WELL_MODE].value = props["well_mode"].item()
    ws[_ANALYTICS_WELLBORE].value = props["wellbore"].item()
    ws[_ANALYTICS_WELL_LIFT].value = props["well_lift"].item()
    ws[_ANALYTICS_TOP_PERF_DEPTH].value = props["top_perf"].item()
    ws[_ANALYTICS_TOP_PERF_OFFSET].value = props["top_perf_offset"].item()
    ws[_ANALYTICS_WATERCUT].value = props["watercut"].item()
    ws[_ANALYTICS_WELL_STATUS].value = props["well_status"].item()
    ws[_ANALYTICS_OIL_DENSITY].value = props["layer_oil_density"].item()
    ws[_ANALYTICS_WATER_DENSITY].value = props["water_density"].item()
    ws[_ANALYTICS_WELL_TEST_START_DATE].value = on_date
    ws[_ANALYTICS_WELL_TEST_END_DATE].value = on_date
    ws[_ANALYTICS_TOP_PERF_PRESSURE].value = props["top_perf_pressure"].item()
    ws[_ANALYTICS_OWC_PRESSURE].value = props["owc_pressure"].item()
    if well_test is WellTest.static_level:
        ws[_ANALYTICS_WELL_TEST_PLAN].value = "Pпл по Hст"
        ws[_ANALYTICS_WELL_TEST_ACTUAL].value = "Hст"
        ws[_ANALYTICS_STATIC_LEVEL_DEPTH].value = props["pressure"].item()
        ws[_ANALYTICS_ANNULAR_PRESSURE].value = props["depth"].item()
    elif well_test is WellTest.pressure:
        ws[_ANALYTICS_WELL_TEST_PLAN].value = "Pпл"
        ws[_ANALYTICS_WELL_TEST_ACTUAL].value = "Pпл"
        ws[_ANALYTICS_MEASURED_DEPTH].value = props["pressure"].item()
        ws[_ANALYTICS_MEASURED_PRESSURE].value = props["depth"].item()


def _process_calculator(
    dfs: dict[str, pd.DataFrame], path: Path, template: Path
):
    result = path / template.name
    try:
        wb = openpyxl.load_workbook(template)
        _fill_calculator(wb["Пересчет"], dfs["props"])
        _fill_depth(wb["Глубины"], dfs["depths"])
        save_workbook(wb, result)
    finally:
        wb.close()


def _process_analytics(
    dfs: dict[str, pd.DataFrame],
    well_test: WellTest,
    path: Path,
    template: Path,
):
    result = path / template.name
    try:
        wb = openpyxl.load_workbook(template)
        _fill_analytics(wb["Лист1"], dfs["props"], well_test)
        save_workbook(wb, result)
    finally:
        wb.close()


def _calc_pressures(
    props: pd.DataFrame, depths: pd.DataFrame, pressure: float, depth: float
) -> None:
    props[["pressure", "depth"]] = pressure, depth
    props[["depth_offset", "top_perf_offset"]] = np.interp(
        (depth, props["top_perf"].item()), depths["md"], depths["offset"]
    )
    props["liquid_density"] = (
        props["layer_oil_density"] * (1 - props["watercut"] / 100)
        + props["water_density"] * props["watercut"] / 100
    )
    props["top_perf_pressure"] = (
        props["pressure"]
        + (
            (props["top_perf"] - props["top_perf_offset"])
            - (props["depth"] - props["depth_offset"])
        )
        * props["liquid_density"]
        / 10
    )
    props["owc_pressure"] = (
        props["pressure"]
        + (
            props["abs_depth_owc"]
            - (props["depth"] - props["depth_offset"] - props["elevation"])
        )
        * props["liquid_density"]
        / 10
    )


def _validate_inputs(dfs: dict[str, pd.DataFrame], well: str) -> None:
    if len(dfs["props"]) != 1:
        raise ValueError(f"Отчет по ВНК не нашел параметры по скважине {well}")
    if dfs["depths"].empty:
        raise ValueError(
            f"Отчет по ВНК не нашел инклинометрию по скважине {well}"
        )


async def owc_resp_report(
    path: Path,
    calculator_template: Path,
    analytics_template: Path,
    field: UneftFieldDB,
    reservoir: UneftReservoirDB,
    well: str,
    pressure: float,
    depth: float,
    well_test: WellTest,
    on_date: date,
    dao: OwcRespReporter,
    pool: ProcessPoolManager,
) -> None:
    dfs = await dao.read_all(
        field_id=field.id,
        reservoir_id=reservoir.id,
        well=well,
        on_date=on_date,
    )
    _validate_inputs(dfs, well)
    _calc_pressures(dfs["props"], dfs["depths"], pressure, depth)
    async with asyncio.TaskGroup() as tg:
        tg.create_task(
            pool.run(_process_calculator, dfs, path, calculator_template)
        )
        tg.create_task(
            pool.run(
                _process_analytics, dfs, well_test, path, analytics_template
            )
        )
    make_archive(str(path), "zip", root_dir=path)
