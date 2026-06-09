import logging
from collections.abc import Awaitable, Callable
from datetime import date
from filecmp import cmpfiles
from pathlib import Path

import pytest
import structlog
from csv_diff import compare, load_csv
from openpyxl import load_workbook

from app.common.config.models.paths import Paths
from app.core.models.dto import UneftFieldDB
from app.core.services.reports import (
    fnv_report,
    inj_loss_report,
    matrix_report,
    oil_loss_report,
    opp_per_year_report,
    owc_resp_report,
    profile_report,
)
from app.core.utils.process_pool import ProcessPoolManager
from app.infrastructure.db.dao.sql.reporters import LocalBaseDAO
from app.infrastructure.files.config.models.csv import CsvSettings
from app.infrastructure.holder import HolderDAO
from tests.fixtures.task_fixtures import (  # noqa
    task_owc_resp_pressure,
    task_owc_resp_static,
)


@pytest.mark.parametrize(
    "dao,service,expected_report,resulted_report",
    [
        (
            "well_profile_reporter",
            profile_report,
            "profile_report.csv",
            "profile.csv",
        ),
        (
            "first_rate_inj_loss_reporter",
            inj_loss_report,
            "first_rate_inj_loss_report.csv",
            "inj_loss.csv",
        ),
        (
            "max_rate_inj_loss_reporter",
            inj_loss_report,
            "max_rate_inj_loss_report.csv",
            "inj_loss.csv",
        ),
        (
            "first_rate_oil_loss_reporter",
            oil_loss_report,
            "first_rate_oil_loss_report.csv",
            "oil_loss.csv",
        ),
        (
            "max_rate_oil_loss_reporter",
            oil_loss_report,
            "max_rate_oil_loss_report.csv",
            "oil_loss.csv",
        ),
        (
            "opp_per_year_reporter",
            opp_per_year_report,
            "opp_per_year_report.csv",
            "opp_per_year.csv",
        ),
        (
            "matrix_reporter",
            matrix_report,
            "matrix_report.csv",
            "matrix.csv",
        ),
    ],
)
@pytest.mark.asyncio(scope="session")
async def test_reports(
    pool_holder: HolderDAO,
    process_pool: ProcessPoolManager,
    tmp_path: Path,
    dao: str,
    service: Callable[..., Awaitable],
    expected_report: str,
    resulted_report: str,
    result_dir: Path,
):
    date_from = date(2000, 1, 1)
    date_to = date(2001, 1, 1)
    dao_: LocalBaseDAO = getattr(pool_holder, dao)
    csv_config = CsvSettings()
    if dao == "matrix_reporter":
        await service(
            tmp_path,
            date_from,
            date_to,
            1,
            12,
            [],
            date_to,
            dao_,
            process_pool,
            ",",
            csv_config,
        )
    elif dao in (
        "opp_per_year_reporter",
        "first_rate_oil_loss_reporter",
        "max_rate_oil_loss_reporter",
    ):
        await service(
            tmp_path, date_from, date_to, dao_, process_pool, csv_config
        )
    elif dao in (
        "first_rate_inj_loss_reporter",
        "max_rate_inj_loss_reporter",
    ):
        await service(
            tmp_path,
            date_from,
            date_to,
            True,
            dao_,
            process_pool,
            ",",
            csv_config,
        )
    else:
        await service(
            tmp_path, date_from, date_to, dao_, process_pool, ",", csv_config
        )
    diff = compare(
        load_csv(open(result_dir / expected_report)),
        load_csv(open(tmp_path / resulted_report)),
    )
    for value in diff.values():
        assert not value


@pytest.mark.asyncio(scope="session")
async def test_fnv_report(
    pool_holder: HolderDAO,
    tmp_path: Path,
    result_dir: Path,
    caplog,
):
    caplog.set_level(logging.DEBUG)
    structlog.contextvars.clear_contextvars()
    structlog.contextvars.bind_contextvars(
        request_id="1234567890", user_id="test_user"
    )
    fnv_dir = result_dir / "fnv"
    fields = [UneftFieldDB(id=1, name="F1")]
    min_radius = 0
    alternative = False
    max_fields = 1
    await fnv_report(
        tmp_path,
        fields,
        min_radius,
        alternative,
        max_fields,
        pool_holder.fnv_reporter,
    )
    parts = ["/".join(file.parts[-2:]) for file in fnv_dir.glob("*/*.txt")]
    _, mismatch, errors = cmpfiles(tmp_path, fnv_dir, parts)
    assert not mismatch
    assert not errors


@pytest.mark.asyncio(scope="session")
@pytest.mark.parametrize(
    (
        "task_fixture,expected_plan,expected_bl,expected_bm,"
        "expected_bn,expected_bo"
    ),
    [
        (
            "task_owc_resp_static",
            "Pпл по Hст",
            100,
            75,
            None,
            None,
        ),
        (
            "task_owc_resp_pressure",
            "Pпл",
            None,
            None,
            100,
            75,
        ),
    ],
)
async def test_owc_resp_report(
    pool_holder: HolderDAO,
    process_pool: ProcessPoolManager,
    paths: Paths,
    tmp_path: Path,
    request,
    task_fixture: str,
    expected_plan: str,
    expected_bl: float | None,
    expected_bm: float | None,
    expected_bn: float | None,
    expected_bo: float | None,
):
    task = request.getfixturevalue(task_fixture)
    path = tmp_path / "owc_resp"
    path.mkdir()
    data_dir = paths.base_dir / "data"
    calculator_template = data_dir / "owc_resp_template.xlsx"
    analytics_template = data_dir / "analytics_template.xlsx"

    await owc_resp_report(
        path,
        calculator_template,
        analytics_template,
        task.field,
        task.reservoir,
        task.well,
        task.pressure,
        task.depth,
        task.well_test,
        task.on_date,
        pool_holder.owc_resp_reporter,
        process_pool,
    )

    analytics = load_workbook(
        path / "analytics_template.xlsx", data_only=False
    )
    calc = load_workbook(path / "owc_resp_template.xlsx", data_only=False)

    analytics_ws = analytics["Лист1"]
    calc_ws = calc["Пересчет"]

    assert analytics_ws["AW5"].value == expected_plan
    assert analytics_ws["BE5"].value == "20.10.2025"
    assert analytics_ws["BF5"].value == "20.10.2025"
    assert analytics_ws["BL5"].value == expected_bl
    assert analytics_ws["BM5"].value == expected_bm
    assert analytics_ws["BN5"].value == expected_bn
    assert analytics_ws["BO5"].value == expected_bo
    assert analytics_ws["BY5"].value == pytest.approx(102.5)
    assert analytics_ws["BZ5"].value == pytest.approx(103.5)
    assert calc_ws["B15"].value == pytest.approx(1.0)
